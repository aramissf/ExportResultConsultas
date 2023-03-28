# importando bibliotecas
import sys
from design import Ui_MainWindow
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox, QFileDialog
import pyodbc
from openpyxl import Workbook
from tkinter import messagebox
import csv
import pandas as pd
import os
import	configparser
from tqdm import tqdm
import time

class Exporta(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        #Chama o metodo que printa as le as configs e printa
        self.leconfi()
        
        
        #salva as configurações
        self.btn_salva_conf.clicked.connect(self.salvaConfig)

        #Seleciona diretorio salvar arquivo
        self.btn_busca_cam.clicked.connect(self.abre_pasta)

        #Exporta a consulta
        self.btn_exportar.clicked.connect(self.executaConsulta)
    
    def leconfi (self):
        if os.path.isfile('config.ini'):

            #Criando o objeto do configparser
            config = configparser.ConfigParser()
            
            #Lendo o arquivo 
            config.read('config.ini')

            #Lendo os parametros no arquivo
            valorServ = config.get('DEFAULT', 'SERVIDOR')
            valorBd = config.get('DEFAULT','BANCO')
            valorUsr = config.get('DEFAULT','USUARIO')
            valorSen = config.get('DEFAULT','SENHA')
            self.tx_servidor.setText(valorServ)
            self.tx_banco.setText(valorBd)
            self.tx_user.setText(valorUsr)
            self.tx_senha.setText(valorSen)

    #Criando função para selecionar diretorio de salvar o arquivo
    def abre_pasta (self):

        diretorio = QFileDialog.getExistingDirectory(self, "Selecione o diretorio")
        
        #verificando se o diretorio foi selecionar e demonstrando na barra
        if diretorio:
            self.tx_local_salvar.setText(diretorio)

    def salvaConfig(self):
        #Lendo os parametro de configuração preenchidos
        servidor = self.tx_servidor.text()
        banco = self.tx_banco.text()
        usuario = self.tx_user.text()
        senha = self.tx_senha.text() 
        nome = self.tx_nome_arquivo.text()
        localsalvar = self.tx_local_salvar.text()
        
        #Criando arquivo de configuracao
        config = configparser.ConfigParser()

        config['DEFAULT'] = {
            'SERVIDOR': servidor,
            'BANCO' : banco,
            'USUARIO' : usuario,
            'SENHA' : senha,
            'CAMINHO': localsalvar}
        
        
        with open('config.ini','w') as configfile:
          
            config.write(configfile)
        messagebox.showinfo("Messagem", "Salvo com Sucesso!")


    def executaConsulta(self):
        try:
            if self.btn_xlsx.isChecked():   
                from openpyxl import Workbook
                #criando string de conexão
                DRIVER = '{ODBC Driver 18 for SQL Server}'
                SERVER = self.tx_servidor.text() 
                DATABASE = self.tx_banco.text()
                UID = self.tx_user.text()
                PWD = self.tx_senha.text()
                NomeArq = self.tx_nome_arquivo.text()
                Diretorio = self.tx_local_salvar.text()
                
                dados_con = 'DRIVER='+DRIVER+";"+'SERVER='+SERVER+";"+'DATABASE='+DATABASE+";"+"Encrypt=no"+";"+'UID='+UID+";"+'Pwd='+PWD+";"
                dados_conexao = dados_con
                print(dados_conexao)
                #Criando conexão 
                conexao = pyodbc.connect(dados_conexao)
                cursor = conexao.cursor()
                consulta = self.textEdit.toPlainText()
                #Executando consulta 
                cursor.execute(f"{consulta}")
                df = cursor.fetchall()
                #exportando a consulta
                Workbook = Workbook()
                #criando a planilha
                worksheet = Workbook.active
                messagebox.showinfo("Messagem", "Processamento iniciado, isso pode demorar um pouco")
                #Escrevendo o cabeçalho na primeira linha
                for i, desc in enumerate(cursor.description):
                    worksheet.cell(row=1, column=i+1, value=desc[0])

                #Barra de progresso para geracao do arquivo
                for i in tqdm(range(100)):
                    self.progressBar.setValue(i)
                    QApplication.processEvents()
                    #Escrevendo os dados na celulas
                    for r, row, in enumerate(df):
                        for c, col in enumerate(row):
                            worksheet.cell(row=r+2, column=c+1, value=str(col))
                self.progressBar.setValue(0)
                formato=(str(NomeArq)+'.xlsx')


                #Exportando a planilha
                caminhoSalvar = Diretorio + "\\" + formato
                Workbook.save(filename=f"{caminhoSalvar}")

                #Fechando a conexão 
                cursor.close()
                conexao.close()

                messagebox.showinfo("Messagem", "Exportado com Sucesso!")


            if self.btn_csv.isChecked():
                #criando string de conexão
                DRIVER = '{ODBC Driver 18 for SQL Server}'
                SERVER = self.tx_servidor.text() 
                DATABASE = self.tx_banco.text()
                UID = self.tx_user.text()
                PWD = self.tx_senha.text()
                NomeArq = self.tx_nome_arquivo.text()
                Diretorio = self.tx_local_salvar.text()
                
                dados_con = 'DRIVER='+DRIVER+";"+'SERVER='+SERVER+";"+'DATABASE='+DATABASE+";"+"Encrypt=no"+";"+'UID='+UID+";"+'Pwd='+PWD+";"
                dados_conexao = dados_con
                
                #Criando conexão 
                conexao = pyodbc.connect(dados_conexao)
                cursor = conexao.cursor()
                consulta = self.textEdit.toPlainText()
                messagebox.showinfo("Messagem", "Processamento iniciado, isso pode demorar um pouco")
                df = pd.read_sql(consulta, conexao)
                tab = pd.DataFrame()

                
                for i, row in tqdm(df.iterrows(), total=len(df)):                  
                    tab = tab.append(row, ignore_index=True)

                    progess =int((i+1)/len(df)*100)
                    self.progressBar.setValue(progess)
                    QApplication.processEvents()
                caminhoSalvar = Diretorio + "\\" + NomeArq+'.csv'
                tab.to_csv(caminhoSalvar, index=False)
                self.progressBar.setValue(0)
                #Fechando a conexão 
                cursor.close()
                conexao.close()              

                messagebox.showinfo("Messagem", "Exportado com Sucesso!")

        except:
            messagebox.showinfo("Messagem", "Erro de conexão, verifique as configurações")

        
if __name__ == '__main__':
    qt = QApplication(sys.argv)
    exporta = Exporta()
    exporta.show()
    qt.exec_()








