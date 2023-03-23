# importando bibliotecas
import pyodbc
from openpyxl import Workbook
#criando string de conexão

dados_conexao = 'DRIVER={ODBC Driver 18 for SQL Server};SERVER=GLOBALTEC-SQL\BI;DATABASE=UAU-BI-SUP;Encrypt=no;UID=DESENVOLVIMENT@;Pwd=7DES@#$uau!7;'

 #'DRIVER={SQL Server};SERVER=Globaltec-sql\bi;DATABASE=UAU-BI-SUP;Trusted_connection=yes' 

#Criando conexão 

conexao = pyodbc.connect(dados_conexao)
#print("conexao foi")

cursor = conexao.cursor()
#Executando consulta 
cursor.execute("SELECT * FROM Empresas")
df = cursor.fetchall()

#exportando a consulta

Workbook = Workbook()
#criando a planilha
worksheet = Workbook.active

#Escrevendo o cabeçalho na primeira linha
for i, desc in enumerate(cursor.description):
    worksheet.cell(row=1, column=i+1, value=desc[0])

#Escrevendo os dados na celulas
for r, row, in enumerate(df):
    for c, col in enumerate(row):
        worksheet.cell(row=r+2, column=c+1, value=str(col))

#Exportando a planilha
Workbook.save(filename="empresas.csv")

#Fechando a conexão 
cursor.close()
conexao.close()